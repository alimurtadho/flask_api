import os, sys
import numpy
import pandas
import itertools
import time
import logging
from sklearn import datasets, metrics, model_selection, preprocessing, linear_model, svm, naive_bayes, neighbors, neural_network, ensemble
from sklearn.externals import joblib
from scipy import stats
import pickle
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import openpyxl as xl

def LoadDataset(dataset_file):
    # Load Data
    dataframe = pandas.read_csv(dataset_file)
    num_rows = dataframe.shape[0]
    logging.info("%d Rows in Dataset" % num_rows)
    return dataframe


def ProcessData(dataframe, options):

    if 'drop_na' in options and options['drop_na']:
        dataframe = dataframe.dropna()
        num_rows = dataframe.shape[0]
        logging.info("%d Rows in Dataset after removing null values" % num_rows)

    if 'replace_zero_values' in options:
        for field in options['replace_zero_values']:
            impute_zero_field(dataframe, field)

    features = dataframe.ix[:,:-1].values
    standard_scalar = preprocessing.StandardScaler().fit(features)
    features_std = standard_scalar.transform(features)

    predictions = dataframe.iloc[:,-1].values

    return (standard_scalar, features_std, predictions)


def impute_zero_field(data, field):
    nonzero_vals = data.loc[data[field] != 0, field]
    avg = numpy.sum(nonzero_vals) / len(nonzero_vals)
    k = len(data.loc[ data[field] == 0, field])   # num of 0-entries
    data.loc[ data[field] == 0, field ] = avg
    logging.info('Field: %s; fixed %d entries with value: %.3f' % (field, k, avg))


def GenerateNeurons(number_of_layers=1, min_neurons=3, max_neurons=50):
    layers = []
    for i in range(min_neurons, max_neurons+1):
        layers.append((i,))
    if number_of_layers > 1:
        items = [list(range(min_neurons, max_neurons+1))] * number_of_layers
        layers += itertools.product(*items)
    return layers


def main():
    logging.basicConfig(filename='log',level=logging.INFO, format='%(asctime)s %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p')
    numpy.random.seed(0)

    options = dict(
        cv=10,
        scoring=['accuracy', 'average_precision', 'f1', 'f1_micro', 'f1_macro', 'f1_weighted', 'neg_log_loss', 'precision', 'recall', 'roc_auc'],
        verbose=2,
        logbook=dict(
            log=True,
            filename="results/Log-%d.xlsx" % (time.time()),
            scores_worksheet="Scores",
            optimizations_worksheet="Optimizations",
            estimator_file="models/%s-%s-%s.pkl"
        )
    )

    logbook = xl.Workbook()
    scores_worksheet = logbook.create_sheet(title=options['logbook']['scores_worksheet'])
    optimizations_worksheet = logbook.create_sheet(title=options['logbook']['optimizations_worksheet'])

    scores_worksheet.append(['Time', 'Dataset', 'Classifier', 'Scoring', 'Scores', 'Mean Score', 'Std Dev'])
    optimizations_worksheet.append(['Time', 'Dataset', 'Classifier', 'Scoring', 'Best Score', 'Best Params', 'Best Estimator', 'File'])

    datasets = {
        'Heart Disease': {
            'filename': 'datasets/cleveland.csv',
            'drop_na': True,
        },
        'Diabetes': {
            'filename': 'datasets/diabetes.csv',
            'replace_zero_values': ['Glucose', 'BloodPressure', 'SkinThickness', 'Insulin', 'BMI']
        },
    }

    classifiers = {
        'Logistic Regression': {
            'clf': linear_model.LogisticRegression(random_state=37,C=0.13, penalty='l1'),
            'cv_params': dict(C=numpy.arange(0.1,5,0.01).tolist(), penalty=['l1', 'l2']),
            'optimize': True,
            'random': False
        },

        'Linear SVC': {
            'clf': svm.LinearSVC(random_state=37,C=18.09),
            'cv_params': dict(C=numpy.arange(0.1,50,0.01).tolist()),
            'optimize': True,
            'random': False
        },

        'Naive Bayes': {
            'clf': naive_bayes.GaussianNB(),
            'cv_params': dict(),
            'optimize': True,
            'random': False
        },

        'K-Nearest Neighbors': {
            'clf': neighbors.KNeighborsClassifier(algorithm='brute', n_jobs=-1, n_neighbors=13, weights='uniform'),
            'cv_params': dict(n_neighbors=numpy.arange(1,50).tolist(), weights=['uniform', 'distance']),
            'optimize': True,
            'random': False
        },

        'MLP Classifier': {
            'clf': neural_network.MLPClassifier(random_state=37,learning_rate_init=0.026958815931057856, learning_rate='constant', hidden_layer_sizes=(29,26,5), activation='identity', alpha=16.681005372000556,max_iter=5000),
            'cv_params': dict(
                hidden_layer_sizes=GenerateNeurons(2),
                activation=['identity', 'logistic', 'tanh', 'relu'],
                learning_rate=['constant', 'invscaling', 'adaptive'],
                alpha=numpy.logspace(-5, 3, 5),learning_rate_init=stats.uniform(0.001, 0.05)
            ),
            'optimize': True,
            'random': True,
            'random_iterations': 10000
        },

    }

    try:
        for dataset_name, dataset_options in datasets.items():

            logging.info("Dataset: %s" % (dataset_name))

            dataframe = LoadDataset(dataset_options['filename'])
            standard_scalar, features, predictions = ProcessData(dataframe, dataset_options)

            for classifier_name, classifier in classifiers.items():
                logging.info("-- Processing: %s ---" % (classifier_name))

                if not classifier['optimize']:
                    scoring_options = options['scoring'] if type(options['scoring']) is list else [options['scoring']]
                    for scoring in scoring_options:
                        if scoring is 'neg_log_loss' and not hasattr(classifier['clf'], 'predict_proba'):
                            logging.info("--- --- Skipping %s for %s" % (scoring, classifier_name))
                            continue
                        scores = model_selection.cross_val_score(classifier['clf'], features, predictions, cv=options['cv'], scoring=scoring)
                        scores_worksheet.append([time.time(), dataset_name, classifier_name, scoring, str(scores), scores.mean(), scores.std()])
                        logging.info("--- --- %.2f%% %s" % (scores.mean() * 100, scoring))
                else:
                    logging.info("--- --- Optimizing Hyper-Parameters")

                    scoring = options['scoring'][0] if type(options['scoring']) is list else options['scoring']
                    grid = model_selection.GridSearchCV(classifier['clf'], classifier['cv_params'], cv=options['cv'], scoring=scoring, n_jobs=-1, verbose=options['verbose']) if not classifier['random'] else model_selection.RandomizedSearchCV(classifier['clf'], classifier['cv_params'], n_iter=classifier['random_iterations'], cv=options['cv'], scoring=scoring, n_jobs=-1, verbose=options['verbose'])

                    grid.fit(features, predictions)

                    dump_file = options['logbook']['estimator_file'] % (dataset_name, classifier_name, time.time())
                    joblib.dump(grid.best_estimator_, dump_file);

                    optimizations_worksheet.append([time.time(), dataset_name, classifier_name, scoring, grid.best_score_, str(grid.best_params_), str(grid.best_estimator_), dump_file])

                    logging.info("--- --- %.2f%% %s with %s" % (grid.best_score_ * 100, scoring, classifier_name))
                    for key, value in grid.best_params_.items(): logging.info("--- --- -- %s: %s" % (key, value))

                logging.info("--- Finished Processing")
    except KeyboardInterrupt:
        logging.error("XXXXX Keyboard Interrupt XXXXX")
    finally:
        logging.info("Saving Worksheet %s" %(options['logbook']['filename']))
        logbook.save(filename = options['logbook']['filename'])

if __name__ == '__main__':
    main()
